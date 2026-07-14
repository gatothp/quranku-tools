"""
SCRIPT: 5-match-ayah.py
PURPOSE: Automatically detect circle markers on Quran pages and match them to ayat (verse) numbers

INPUT:
  - User: Page number or range (e.g., "3" or "3-5")
  - Image: Quran mushaf pages (jpg) from madinah_v2 folder
  - Template: circle.png - template image for circle detection
  - Excel: quran.xlsx - contains ayat ranges and page mappings

METHOD:
  1. Load Excel file to get ayat ranges for the page
  2. Read Quran page image and convert to grayscale
  3. Use template matching to detect circle markers at multiple scales (0.8-1.2x)
  4. Filter duplicates based on proximity (20px threshold)
  5. Expand bounding boxes by 40% for full circle coverage
  6. Sort detected circles: top-to-bottom, right-to-left within each line
  7. Assign ayat numbers sequentially based on sorted position
  8. Validate count against Excel data

OUTPUT:
  - JSON: bbox/p###.json - Contains page number and detected circles with:
    * id, x, y, width, height (bounding box)
    * surat (chapter), ayat (verse) numbers
  - Console: Detection summary and validation status
"""

import cv2
import numpy as np
import json
import os
import openpyxl

# Create bbox directory if it doesn't exist
os.makedirs("bbox", exist_ok=True)

# Prompt for page number or range
page_input = input("Enter page number or range (e.g., 7 or 3-4): ").strip()

# Parse page input
if "-" in page_input:
    try:
        start_page, end_page = map(int, page_input.split("-"))
        page_range = list(range(start_page, end_page + 1))
    except:
        print("❌ Invalid range format. Use 'N' or 'N-M'")
        exit(1)
else:
    try:
        page_num = int(page_input)
        page_range = [page_num]
    except:
        print("❌ Invalid page number")
        exit(1)

def extract_number_from_circle(gray_img, x, y, w, h):
    """Values come from Excel file, not OCR"""
    return ""

# Process each page in the range
for page_num in page_range:
    IMAGE_PATH = f"D:\\apps\\android\\quran\\app\\src\\main\\assets\\files\\mushaf\\madinah_v2\\p{page_num:03d}.jpg"
    TEMPLATE_PATH = "files/circle.png"
    OUTPUT_JSON = f"bbox/p{page_num:03d}.json"
    EXCEL_PATH = "files/quran.xlsx"
    
    # Load surah mapping from Excel
    surah_map = {}  # ayat_number -> surat_number
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb['Quran Pages']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[4] == page_num:  # Column E is page (index 4)
                surat = row[0]  # Column A is surah (index 0)
                start_ayat = row[2]  # Column C is start ayat (index 2)
                end_ayat = row[3]  # Column D is end ayat (index 3)
                
                if start_ayat and end_ayat:
                    for ayat_num in range(int(start_ayat), int(end_ayat) + 1):
                        surah_map[ayat_num] = surat
        # print(f"✅ Loaded surah mapping: {surah_map}")
    except Exception as e:
        print(f"⚠️  Could not load surah mapping from Excel: {e}")
    
    img = cv2.imread(IMAGE_PATH)
    if img is None:
        print(f"❌ Image not found: {IMAGE_PATH}")
        continue
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    
    template = cv2.imread(TEMPLATE_PATH, 0)
    if template is None:
        print(f"❌ Template not found: {TEMPLATE_PATH}")
        continue
    
    tw, th = template.shape[::-1]
    
    # Try multiple scales
    scales = np.linspace(0.8, 1.2, 10)
    
    results = []
    seen = []
    
    # Get ayat range for this page from Excel
    ayat_range = []
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb['Quran Pages']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[4] == page_num:  # Column E is page (index 4)
                start_ayat = row[2]  # Column C is start ayat (index 2)
                end_ayat = row[3]  # Column D is end ayat (index 3)
                surat = row[0]  # Column A is surah (index 0)
                
                if start_ayat and end_ayat:
                    ayat_range.append((surat, int(start_ayat), int(end_ayat)))
        
        print(f"✅ Page {page_num} ayat ranges: {ayat_range}")
    except Exception as e:
        print(f"⚠️  Could not load ayat ranges from Excel: {e}")
    
    # Flatten ayat_range into a list of (ayat_num, surat_num) tuples, sorted by position
    all_ayats = []
    for surat, start, end in ayat_range:
        for ayat_num in range(start, end + 1):
            all_ayats.append((ayat_num, surat))
    
    ayat_index = 0  # Track which ayat we're on
    
    for scale in scales:
        resized = cv2.resize(gray, None, fx=scale, fy=scale)
        result = cv2.matchTemplate(resized, template, cv2.TM_CCOEFF_NORMED)
        
        threshold = 0.3
        loc = np.where(result >= threshold)
        
        for pt in zip(*loc[::-1]):
            # Convert back to original scale
            x = int(pt[0] / scale)
            y = int(pt[1] / scale)
            w = int(tw / scale)
            h = int(th / scale)
            
            # Expand bounding box by 40%
            expand = 1.4
            x = int(x - w * (expand - 1) / 2)
            y = int(y - h * (expand - 1) / 2)
            w = int(w * expand)
            h = int(h * expand)
            
            # Avoid duplicates
            too_close = False
            for (sx, sy) in seen:
                if abs(x - sx) < 20 and abs(y - sy) < 20:
                    too_close = True
                    break
            
            if too_close:
                continue
            
            seen.append((x, y))
            
            results.append({
                "id": len(results) + 1,
                "x": x,
                "y": y,
                "width": w,
                "height": h,
                "surat": 0,
                "ayat": 0
            })
    
    # Sort results: top to bottom, but within each line sort right to left
    # Group boxes by Y-coordinate (same line) with tolerance
    def get_line_group(box):
        # Group boxes into lines (every ~60-80 pixels in height)
        return box["y"] // 80
    
    # First sort by line, then by X position (right to left)
    results.sort(key=lambda box: (get_line_group(box), -box["x"]))
    
    # Assign ayat numbers based on sorted order
    for i, result in enumerate(results):
        if i < len(all_ayats):
            ayat_num, surat_num = all_ayats[i]
            result["ayat"] = ayat_num
            result["surat"] = surat_num
        else:
            result["ayat"] = 0
            result["surat"] = 0
    
    # Assign IDs after sorting
    for idx, result in enumerate(results, 1):
        result["id"] = idx
    
    # Save results
    with open(OUTPUT_JSON, "w") as f:
        json.dump({"page": page_num, "boxes": results}, f, ensure_ascii=False, indent=2)
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb['hal-ayat']
        expected_ayat = None
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == page_num:
                expected_ayat = row[1]
                break
    except Exception as e:
        print(f"⚠️  Could not load Excel file: {e}")
        expected_ayat = None
    
    if expected_ayat:
        if len(results) == expected_ayat:
            print(f"✅ Validation: {len(results)} detected (matches {expected_ayat} ayat)")
        else:
            print(f"⚠️  Validation: {len(results)} detected but {expected_ayat} ayat expected")
    else:
        print(f"ℹ️  {len(results)} circles detected")
    
print(f"\n✅ Done!")
